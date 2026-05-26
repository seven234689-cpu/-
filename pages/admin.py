from dash import html, dcc, Input, Output, State, dash_table, ctx, no_update
import json
import base64, io, pandas as pd
import sqlalchemy as sa
import db

LAO = {'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif','fontSize':'13px'}

def ok_box(msg, detail=None):
    return html.Div(style={'background':'#E8F5E9','borderRadius':'10px',
        'border':'1px solid #A5D6A7','padding':'12px 16px'}, children=[
        html.Div(style={'display':'flex','alignItems':'center','gap':'8px'}, children=[
            html.Span('✓', style={'color':db.GREEN,'fontWeight':'700','fontSize':'16px'}),
            html.Span(msg, style={**LAO,'color':'#1B5E20','fontWeight':'600'})
        ]),
        html.Div(detail, style={**LAO,'color':'#388E3C','marginTop':'4px'}) if detail else None
    ])

def err_box(msg):
    return html.Div(msg, style={**LAO,'color':'#B71C1C','background':'#FFEBEE',
        'padding':'10px 14px','borderRadius':'8px','border':'1px solid #FFCDD2'})

def warn_box(msg):
    return html.Div(msg, style={**LAO,'color':'#E65100','background':'#FFF3E0',
        'padding':'10px 14px','borderRadius':'8px','border':'1px solid #FFCC80'})

def inp_s():
    return {'width':'100%','padding':'10px 14px','fontSize':'13px',
            'border':f'1.5px solid {db.BD}','borderRadius':'8px','outline':'none',
            'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif',
            'color':db.TX2,'background':'white'}

def lbl(t):
    return html.Label(t, style={**LAO,'fontWeight':'600','color':db.TX,
        'display':'block','marginBottom':'6px','fontSize':'12px'})

def btn(t, bid, color):
    return html.Button(t, id=bid, n_clicks=0, style={
        'padding':'10px 24px','background':color,'color':'white',
        'border':'none','borderRadius':'8px','fontSize':'13px',
        'fontWeight':'600','cursor':'pointer',
        'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'})

def tbl_style():
    return {
        'style_table': {'overflowX':'auto','borderRadius':'10px','overflow':'hidden','border':f'1px solid {db.BD}'},
        'style_header': {'backgroundColor':'#F8FAFD','color':db.TX,'fontWeight':'600',
            'fontSize':'12px','border':'none','padding':'12px 16px','borderBottom':f'1px solid {db.BD}'},
        'style_cell': {'fontSize':'13px','padding':'10px 16px','border':'none',
            'borderBottom':f'1px solid {db.BD}','textAlign':'center',
            'color':db.TX2,'backgroundColor':'white',
            'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'},
        'style_data': {'backgroundColor':'white'},
        'style_data_conditional': [{'if':{'row_index':'odd'},'backgroundColor':'#FAFBFD'}],
        'page_size': 15,
        'sort_action': 'native',
        'filter_action': 'native',
        'style_as_list_view': True,
    }


layout = html.Div(style={'padding':'28px 32px','background':db.PAGE,'minHeight':'100vh'}, children=[

    html.Div(style={'marginBottom':'24px'}, children=[
        html.Div('Admin Panel', style={'fontSize':'22px','fontWeight':'700','color':db.TX2}),
        html.Div('ຈັດການຂໍ້ມູນ · ນຳເຂົ້າ Excel · ເພີ່ມ/ລຶບ ນ.ສ · ແກ້ໄຂເກຣດ · Export',
                 style={'fontSize':'13px','color':db.TX,'marginTop':'4px',
                        'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'}),
    ]),

    # KPI
    html.Div(style={'display':'flex','gap':'12px','marginBottom':'22px','flexWrap':'wrap'}, children=[
        db.kpi_card(db.total,      'ນ.ສ ທັງໝົດ',  db.BLUE),
        db.kpi_card(db.total_subj, 'ວິຊາ',          '#546078'),
        db.kpi_card(db.male,       'ນ.ສ ຊາຍ',       db.BLUE),
        db.kpi_card(db.female,     'ນ.ສ ຍິງ',       '#6A1B9A'),
        db.kpi_card(db.risk,       'ກຸ່ມ ສ່ຽງ',      db.RED),
    ]),

    # ── 1. Upload ────────────────────────────────────────────
    html.Div(style=db.card_style(db.BLUE), children=[
        db.sec_title('📂 ນຳເຂົ້າຂໍ້ມູນຈາກໄຟລ໌ Excel'),
        db.sec_sub('ອັບໂຫລດ .xlsx — ລະບົບຈະເພີ່ມ Student + Score ເຂົ້າ MySQL ອັດຕະໂນມັດ'),
        dcc.Upload(id='adm-upload', accept='.xlsx', multiple=False,
            children=html.Div(style={
                'border':f'2px dashed {db.BD}','borderRadius':'10px',
                'padding':'28px 20px','textAlign':'center',
                'background':db.PAGE,'cursor':'pointer'
            }, children=[
                html.Div('ລາກໄຟລ໌ .xlsx ມາວາງ ຫຼື ກົດເພື່ອເລືອກ',
                         style={**LAO,'color':db.BLUE,'fontWeight':'600'}),
                html.Div('ຂໍ້ມູນທີ່ຊ້ຳຈະຖືກຂ້າມໂດຍອັດຕະໂນມັດ',
                         style={**LAO,'color':'#90A4AE','marginTop':'4px','fontSize':'12px'}),
            ])),
        html.Div(id='adm-upload-result', style={'marginTop':'14px'}),

        # Preview + Confirm/Cancel
        html.Div(id='adm-confirm-section', style={'display':'none'}, children=[
            html.Hr(style={'border':'none','borderTop':f'1px solid {db.BD}','margin':'16px 0'}),
            html.Div(id='adm-preview-content'),
            html.Div(style={'display':'flex','gap':'10px','marginTop':'14px'}, children=[
                html.Button('✓ ຢືນຢັນນຳເຂົ້າ', id='adm-btn-confirm', n_clicks=0, style={
                    'padding':'10px 28px','background':db.GREEN,'color':'white',
                    'border':'none','borderRadius':'8px','fontSize':'13px',
                    'fontWeight':'700','cursor':'pointer',
                    'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'}),
                html.Button('✕ ຍົກເລີກ', id='adm-btn-cancel', n_clicks=0, style={
                    'padding':'10px 24px','background':'#ECEFF1','color':db.TX,
                    'border':f'1px solid {db.BD}','borderRadius':'8px','fontSize':'13px',
                    'fontWeight':'600','cursor':'pointer',
                    'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'}),
            ]),
            html.Div(id='adm-confirm-result', style={'marginTop':'12px'})
        ]),
        dcc.Store(id='adm-pending-data'),
    ]),

    # ── 2. Add Student ───────────────────────────────────────
    html.Div(style=db.card_style(db.GREEN), children=[
        db.sec_title('➕ ເພີ່ມນັກສຶກສາໃໝ່'),
        db.sec_sub('ປ້ອນລະຫັດ ແລະ ເພດ ແລ້ວກົດ "ບັນທຶກ"'),
        html.Div(style={'display':'flex','gap':'12px','flexWrap':'wrap','alignItems':'flex-end'}, children=[
            html.Div(style={'flex':'2','minWidth':'200px'}, children=[
                lbl('ລະຫັດນັກສຶກສາ'),
                dcc.Input(id='adm-code', type='text',
                          placeholder='ເຊັ່ນ 205N0001.25', debounce=False, style=inp_s())
            ]),
            html.Div(style={'flex':'1','minWidth':'160px'}, children=[
                lbl('ເພດ'),
                dcc.Dropdown(id='adm-gender',
                    options=[{'label':'ຊາຍ (Male)','value':'M'},
                              {'label':'ຍິງ (Female)','value':'F'}],
                    placeholder='ເລືອກ...', clearable=False, style={'fontSize':'13px'})
            ]),
            html.Div(style={'paddingBottom':'2px'}, children=[btn('ບັນທຶກ','adm-btn-add',db.GREEN)]),
        ]),
        html.Div(id='adm-add-result', style={'marginTop':'12px'})
    ]),

    # ── 3. Delete Student ────────────────────────────────────
    html.Div(style=db.card_style(db.RED), children=[
        db.sec_title('🗑 ລຶບນັກສຶກສາ'),
        db.sec_sub('ລຶບນ.ສ ແລະ ຄະແນນທັງໝົດຂອງນ.ສ ຄົນນັ້ນອອກຈາກລະບົບ'),
        html.Div(style={'display':'flex','gap':'12px','flexWrap':'wrap','alignItems':'flex-end'}, children=[
            html.Div(style={'flex':'2','minWidth':'200px'}, children=[
                lbl('ລະຫັດນັກສຶກສາທີ່ຕ້ອງການລຶບ'),
                dcc.Input(id='adm-del-code', type='text',
                          placeholder='ເຊັ່ນ 205N0001.25', debounce=False, style=inp_s())
            ]),
            html.Div(style={'paddingBottom':'2px'}, children=[btn('ລຶບ','adm-btn-del',db.RED)]),
        ]),
        html.Div(id='adm-del-result', style={'marginTop':'12px'})
    ]),

    # ── 4. Add Subject ───────────────────────────────────────
    html.Div(style=db.card_style('#6A1B9A'), children=[
        db.sec_title('➕ ເພີ່ມວິຊາໃໝ່'),
        db.sec_sub('ປ້ອນລະຫັດ ແລະ ຊື່ວິຊາ'),
        html.Div(style={'display':'flex','gap':'12px','flexWrap':'wrap','alignItems':'flex-end'}, children=[
            html.Div(style={'flex':'1','minWidth':'150px'}, children=[
                lbl('ລະຫັດວິຊາ'),
                dcc.Input(id='adm-subcode', type='text',
                          placeholder='205CS201', debounce=False, style=inp_s())
            ]),
            html.Div(style={'flex':'3','minWidth':'240px'}, children=[
                lbl('ຊື່ວິຊາ'),
                dcc.Input(id='adm-subname', type='text',
                          placeholder='ຊື່ວິຊາ...', debounce=False, style=inp_s())
            ]),
            html.Div(style={'paddingBottom':'2px'}, children=[btn('ບັນທຶກ','adm-btn-sub','#6A1B9A')]),
        ]),
        html.Div(id='adm-sub-result', style={'marginTop':'12px'})
    ]),

    # ── 5. Delete Subject ────────────────────────────────────
    html.Div(style=db.card_style('#B71C1C'), children=[
        db.sec_title('🗑 ລຶບວິຊາ'),
        db.sec_sub('ລຶບວິຊາ ແລະ ຄະແນນທັງໝົດຂອງວິຊານັ້ນອອກຈາກລະບົບ'),
        html.Div(style={'display':'flex','gap':'12px','flexWrap':'wrap','alignItems':'flex-end'}, children=[
            html.Div(style={'flex':'2','minWidth':'200px'}, children=[
                lbl('ລະຫັດວິຊາທີ່ຕ້ອງການລຶບ'),
                dcc.Input(id='adm-del-subcode', type='text',
                          placeholder='ເຊັ່ນ 205CS201', debounce=False, style=inp_s())
            ]),
            html.Div(style={'paddingBottom':'2px'}, children=[btn('ລຶບ','adm-btn-del-sub','#B71C1C')]),
        ]),
        html.Div(id='adm-del-sub-result', style={'marginTop':'12px'})
    ]),

    # ── 6. Edit Grade ────────────────────────────────────────
    html.Div(style=db.card_style('#E65100'), children=[
        db.sec_title('✏️ ແກ້ໄຂເກຣດ'),
        db.sec_sub('ປ້ອນລະຫັດ ນ.ສ · ລະຫັດວິຊາ · ພາກຮຽນ · ເກຣດໃໝ່'),
        html.Div(style={'display':'flex','gap':'12px','flexWrap':'wrap','alignItems':'flex-end'}, children=[
            html.Div(style={'flex':'2','minWidth':'160px'}, children=[
                lbl('ລະຫັດນັກສຶກສາ'),
                dcc.Input(id='adm-edit-stu', type='text',
                          placeholder='205N0001.25', debounce=False, style=inp_s())
            ]),
            html.Div(style={'flex':'2','minWidth':'130px'}, children=[
                lbl('ລະຫັດວິຊາ'),
                dcc.Input(id='adm-edit-sub', type='text',
                          placeholder='205CS201', debounce=False, style=inp_s())
            ]),
            html.Div(style={'flex':'1','minWidth':'100px'}, children=[
                lbl('ພາກຮຽນ'),
                dcc.Dropdown(id='adm-edit-sem',
                    options=[{'label':s,'value':s} for s in db.sem_order],
                    placeholder='ເລືອກ...', clearable=False, style={'fontSize':'13px'})
            ]),
            html.Div(style={'flex':'1','minWidth':'100px'}, children=[
                lbl('ເກຣດໃໝ່'),
                dcc.Dropdown(id='adm-edit-grade',
                    options=[{'label':g,'value':g} for g in ['A','B+','B','C+','C','D+','D','F']],
                    placeholder='ເລືອກ...', clearable=False, style={'fontSize':'13px'})
            ]),
            html.Div(style={'paddingBottom':'2px'}, children=[btn('ບັນທຶກ','adm-btn-edit','#E65100')]),
        ]),
        html.Div(id='adm-edit-result', style={'marginTop':'12px'})
    ]),

    # ── 7. Export ────────────────────────────────────────────
    html.Div(style=db.card_style('#00695C'), children=[
        db.sec_title('📥 Export ຂໍ້ມູນ'),
        db.sec_sub('ດາວໂຫລດຂໍ້ມູນອອກເປັນໄຟລ໌ Excel'),
        html.Div(style={'display':'flex','gap':'12px','flexWrap':'wrap'}, children=[
            html.A(html.Button('⬇ Export ນັກສຶກສາທັງໝົດ', style={
                'padding':'10px 20px','background':'#00695C','color':'white',
                'border':'none','borderRadius':'8px','fontSize':'13px',
                'fontWeight':'600','cursor':'pointer',
                'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'
            }), href='/export/students', target='_blank'),
            html.A(html.Button('⬇ Export ວິຊາທັງໝົດ', style={
                'padding':'10px 20px','background':'#00838F','color':'white',
                'border':'none','borderRadius':'8px','fontSize':'13px',
                'fontWeight':'600','cursor':'pointer',
                'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'
            }), href='/export/subjects', target='_blank'),
            html.A(html.Button('⬇ Export ຄະແນນທັງໝົດ', style={
                'padding':'10px 20px','background':'#1565C0','color':'white',
                'border':'none','borderRadius':'8px','fontSize':'13px',
                'fontWeight':'600','cursor':'pointer',
                'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'
            }), href='/export/scores', target='_blank'),
        ]),
    ]),

    # ── 8. Reset ─────────────────────────────────────────────
    html.Div(style=db.card_style('#37474F'), children=[
        db.sec_title('⚠️ ລ້າງຂໍ້ມູນທັງໝົດ (Reset)'),
        db.sec_sub('ລຶບຂໍ້ມູນ Student, Subject, Score ທັງໝົດອອກຈາກລະບົບ — ຄືນຄ່າບໍ່ໄດ້!'),
        html.Div(style={'display':'flex','gap':'12px','alignItems':'center','flexWrap':'wrap'}, children=[
            html.Div(style={'flex':'1'}, children=[
                lbl('ພິມ "RESET" ເພື່ອຢືນຢັນ'),
                dcc.Input(id='adm-reset-confirm', type='text',
                          placeholder='RESET', debounce=False, style=inp_s())
            ]),
            html.Div(style={'paddingTop':'20px'}, children=[
                btn('ລ້າງຂໍ້ມູນທັງໝົດ', 'adm-btn-reset', '#37474F')
            ]),
        ]),
        html.Div(id='adm-reset-result', style={'marginTop':'12px'})
    ]),

    # ── 9. Student Table ─────────────────────────────────────
    html.Div(style=db.card_style('#546078'), children=[
        db.sec_title(f'👤 ລາຍຊື່ນ.ສ ທັງໝົດ ({db.total} ຄົນ)'),
        db.sec_sub('ກົດຫົວຖັນເພື່ອຈັດຮຽງ · ຊ່ອງດ້ານລຸ່ມສຳລັບຄົ້ນຫາ'),
        dash_table.DataTable(
            id='adm-stu-table',
            data=db.df_student[['student_id','student_code','gender']].to_dict('records'),
            columns=[{'name':c,'id':c} for c in ['student_id','student_code','gender']],
            style_table={'overflowX':'auto','borderRadius':'10px','overflow':'hidden','border':f'1px solid {db.BD}'},
            style_header={'backgroundColor':'#F8FAFD','color':db.TX,'fontWeight':'600',
                'fontSize':'12px','border':'none','padding':'12px 16px','borderBottom':f'1px solid {db.BD}'},
            style_cell={'fontSize':'13px','padding':'10px 16px','border':'none',
                'borderBottom':f'1px solid {db.BD}','textAlign':'center',
                'color':db.TX2,'backgroundColor':'white',
                'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'},
            style_data={'backgroundColor':'white'},
            style_data_conditional=[
                {'if':{'row_index':'odd'},'backgroundColor':'#FAFBFD'},
                {'if':{'filter_query':'{gender} = "F"'},'color':'#6A1B9A','fontWeight':'600'},
            ],
            page_size=15, sort_action='native', filter_action='native', style_as_list_view=True,
        )
    ]),

    # ── 10. Subject Table ────────────────────────────────────
    html.Div(style=db.card_style(db.BLUE), children=[
        db.sec_title(f'📚 ວິຊາທັງໝົດ ({db.total_subj} ວິຊາ)'),
        db.sec_sub('ກົດຫົວຖັນເພື່ອຈັດຮຽງ · ຊ່ອງດ້ານລຸ່ມສຳລັບຄົ້ນຫາ'),
        dash_table.DataTable(
            id='adm-sub-table',
            data=db.df_subject[['subject_id','subject_code','subject_name']].to_dict('records'),
            columns=[
                {'name':'ID',        'id':'subject_id'},
                {'name':'ລະຫັດວິຊາ', 'id':'subject_code'},
                {'name':'ຊື່ວິຊາ',    'id':'subject_name'},
            ],
            style_table={'overflowX':'auto','borderRadius':'10px','overflow':'hidden','border':f'1px solid {db.BD}'},
            style_header={'backgroundColor':'#F8FAFD','color':db.TX,'fontWeight':'600',
                'fontSize':'12px','border':'none','padding':'12px 16px','borderBottom':f'1px solid {db.BD}'},
            style_cell={'fontSize':'13px','padding':'10px 16px','border':'none',
                'borderBottom':f'1px solid {db.BD}','textAlign':'left',
                'color':db.TX2,'backgroundColor':'white',
                'fontFamily':'Noto Sans Lao,Segoe UI,Arial,sans-serif'},
            style_data={'backgroundColor':'white'},
            style_data_conditional=[{'if':{'row_index':'odd'},'backgroundColor':'#FAFBFD'}],
            style_cell_conditional=[
                {'if':{'column_id':'subject_id'},   'textAlign':'center','width':'60px'},
                {'if':{'column_id':'subject_code'},  'width':'140px'},
                {'if':{'column_id':'subject_name'},  'textAlign':'left'},
            ],
            page_size=15, sort_action='native', filter_action='native', style_as_list_view=True,
        )
    ]),
])


def register_callbacks(app):

    # ── Step 1: วิเคราะห์ไฟล์ (ยังไม่ import จริง) ──────────
    @app.callback(
        Output('adm-upload-result','children'),
        Output('adm-confirm-section','style'),
        Output('adm-preview-content','children'),
        Output('adm-pending-data','data'),
        Input('adm-upload','contents'),
        State('adm-upload','filename'),
        prevent_initial_call=True)
    def analyze_excel(contents, filename):
        hidden = {'display':'none'}
        if not contents: return '', hidden, '', None
        if not filename.endswith('.xlsx'):
            return err_box('ໄຟລ໌ຕ້ອງເປັນ .xlsx'), hidden, '', None
        try:
            import openpyxl
            _, cs = contents.split(',')
            decoded = base64.b64decode(cs)
            wb_in = openpyxl.load_workbook(io.BytesIO(decoded), read_only=True)
            gm = {'A':4.0,'B+':3.5,'B':3.0,'C+':2.5,'C':2.0,'D+':1.5,'D':1.0,'F':0.0}
            new_s=0; new_sc=0; new_sub=0; skip=0; dup_students=[]; dup_subjects=[]
            is_=[]; ib_=[]; ic_=[]
            with db.engine.connect() as conn:
                sm = dict(zip(*[pd.read_sql("SELECT student_code,student_id FROM student",conn)[c]
                                for c in ['student_code','student_id']]))
                bm = dict(zip(*[pd.read_sql("SELECT subject_code,subject_id FROM subject",conn)[c]
                                for c in ['subject_code','subject_id']]))
                ex = pd.read_sql("SELECT student_id,subject_id,semester FROM score",conn)
                sc_set = set(zip(ex['student_id'],ex['subject_id'],ex['semester']))
                ms = int(pd.read_sql("SELECT COALESCE(MAX(student_id),0) AS m FROM student",conn)['m'].iloc[0])
                mb = int(pd.read_sql("SELECT COALESCE(MAX(subject_id),0) AS m FROM subject",conn)['m'].iloc[0])
                mc = int(pd.read_sql("SELECT COALESCE(MAX(score_id),0) AS m FROM score",conn)['m'].iloc[0])
                for sn in wb_in.sheetnames:
                    ws = wb_in[sn]; rows = list(ws.iter_rows(values_only=True))
                    if len(rows)<4: continue
                    sids=[str(v).strip() for v in rows[1][5:] if v is not None]
                    gens=list(rows[0][5:5+len(sids)])
                    for i,sid in enumerate(sids):
                        if sid not in sm:
                            ms+=1; g=str(gens[i]).strip() if i<len(gens) and gens[i] else 'M'
                            gv='F' if g=='ນາງ' else 'M'
                            is_.append({'student_id':ms,'student_code':sid,'gender':gv})
                            sm[sid]=ms; new_s+=1
                        else:
                            if sid not in dup_students: dup_students.append(sid)
                    cur=None
                    for row in rows[3:]:
                        if row[0] is not None: cur=str(row[0]).strip()
                        if row[2] is None: continue
                        sc2=str(row[2]).strip()
                        sn2=str(row[3]).replace('\u200b','').strip() if row[3] else ''
                        if sc2 not in bm:
                            mb+=1; ib_.append({'subject_id':mb,'subject_code':sc2,'subject_name':sn2})
                            bm[sc2]=mb; new_sub+=1
                        else:
                            if sc2 not in dup_subjects: dup_subjects.append(sc2)
                        for i,gr in enumerate(list(row[5:5+len(sids)])):
                            if not gr or str(gr).strip()=='': continue
                            g=str(gr).strip(); gp=gm.get(g)
                            if gp is None: continue
                            si=sm.get(sids[i]); bi=bm.get(sc2)
                            if not si or not bi or not cur: continue
                            key=(si,bi,cur)
                            if key in sc_set: skip+=1; continue
                            mc+=1
                            ic_.append({'score_id':mc,'student_id':si,'subject_id':bi,
                                        'semester':cur,'grade':g,'grade_point':gp})
                            sc_set.add(key); new_sc+=1

            pending = json.dumps({'is_':is_,'ib_':ib_,'ic_':ic_})

            badge = lambda txt,bg,col: html.Span(txt, style={
                **LAO,'background':bg,'color':col,
                'padding':'4px 12px','borderRadius':'6px','fontWeight':'600'})

            preview_blocks = [
                html.Div(f'📋 ສະຫຼຸບໄຟລ໌: {filename}',
                         style={**LAO,'fontWeight':'700','color':db.TX2,'marginBottom':'10px','fontSize':'14px'}),
                html.Div(style={'display':'flex','gap':'8px','flexWrap':'wrap','marginBottom':'12px'}, children=[
                    badge(f'ນ.ສ ໃໝ່: {new_s}',          '#C8E6C9','#1B5E20'),
                    badge(f'ວິຊາໃໝ່: {new_sub}',        '#E8EAF6','#283593'),
                    badge(f'ຄະແນນໃໝ່: {new_sc}',        '#BBDEFB','#0D47A1'),
                    badge(f'ຄະແນນຊ້ຳ (ຂ້າມ): {skip}',  '#ECEFF1',db.TX),
                ]),
            ]

            if dup_students:
                preview_blocks.append(html.Div(style={
                    'background':'#FFF8E1','border':'1px solid #FFE082',
                    'borderRadius':'8px','padding':'10px 14px','marginBottom':'8px'
                }, children=[
                    html.Div('⚠️ ນ.ສ ທີ່ມີຢູ່ແລ້ວໃນລະບົບ (ຈະຂ້າມ)',
                             style={**LAO,'fontWeight':'700','color':'#E65100','marginBottom':'6px'}),
                    html.Div(', '.join(dup_students[:30])+(f' ... ແລະ ອີກ {len(dup_students)-30} ຄົນ' if len(dup_students)>30 else ''),
                             style={**LAO,'color':'#BF360C','fontSize':'12px','lineHeight':'1.8'})
                ]))

            if dup_subjects:
                preview_blocks.append(html.Div(style={
                    'background':'#FFF8E1','border':'1px solid #CE93D8',
                    'borderRadius':'8px','padding':'10px 14px','marginBottom':'8px'
                }, children=[
                    html.Div('⚠️ ວິຊາທີ່ມີຢູ່ແລ້ວໃນລະບົບ (ຈະຂ້າມ)',
                             style={**LAO,'fontWeight':'700','color':'#E65100','marginBottom':'6px'}),
                    html.Div(', '.join(dup_subjects[:30])+(f' ... ແລະ ອີກ {len(dup_subjects)-30} ວິຊາ' if len(dup_subjects)>30 else ''),
                             style={**LAO,'color':'#BF360C','fontSize':'12px','lineHeight':'1.8'})
                ]))

            if skip > 0:
                preview_blocks.append(html.Div(style={
                    'background':'#FFF8E1','border':'1px solid #B0BEC5',
                    'borderRadius':'8px','padding':'10px 14px','marginBottom':'8px'
                }, children=[
                    html.Div(f'⚠️ ຄະແນນຊ້ຳ {skip} ແຖວ — ຈະຖືກຂ້າມທັງໝົດ',
                             style={**LAO,'color':'#E65100','fontSize':'12px'})
                ]))

            upload_msg = html.Div(style={
                'background':'#E3F2FD','border':'1px solid #90CAF9',
                'borderRadius':'8px','padding':'10px 14px'
            }, children=[
                html.Span('📂 ', style={'fontSize':'16px'}),
                html.Span('ວິເຄາະໄຟລ໌ສຳເລັດ — ກວດສອບຂໍ້ມູນດ້ານລຸ່ມ ແລ້ວກົດ "ຢືນຢັນ" ຫຼື "ຍົກເລີກ"',
                          style={**LAO,'color':'#0D47A1','fontWeight':'600'})
            ])

            return upload_msg, {'display':'block'}, html.Div(preview_blocks), pending

        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:120]}'), hidden, '', None

    # ── Step 2a: ยืนยัน import จริง ──────────────────────────
    @app.callback(
        Output('adm-confirm-result','children'),
        Output('adm-confirm-section','style', allow_duplicate=True),
        Output('adm-pending-data','data', allow_duplicate=True),
        Input('adm-btn-confirm','n_clicks'),
        State('adm-pending-data','data'),
        prevent_initial_call=True)
    def confirm_import(n, pending_json):
        if not pending_json: return warn_box('ບໍ່ມີຂໍ້ມູນ'), no_update, no_update
        try:
            pending = json.loads(pending_json)
            is_ = pending['is_']; ib_ = pending['ib_']; ic_ = pending['ic_']
            with db.engine.connect() as conn:
                if is_: conn.execute(sa.text("INSERT INTO student(student_id,student_code,gender) VALUES(:student_id,:student_code,:gender)"),is_)
                if ib_: conn.execute(sa.text("INSERT INTO subject(subject_id,subject_code,subject_name) VALUES(:subject_id,:subject_code,:subject_name)"),ib_)
                if ic_: conn.execute(sa.text("INSERT INTO score(score_id,student_id,subject_id,semester,grade,grade_point) VALUES(:score_id,:student_id,:subject_id,:semester,:grade,:grade_point)"),ic_)
                conn.commit()
            db.reload_data()
            return ok_box('ນຳເຂົ້າສຳເລັດ!',
                          f'ນ.ສ ໃໝ່: {len(is_)} · ວິຊາໃໝ່: {len(ib_)} · ຄະແນນ: {len(ic_)}'),                    {'display':'none'}, None
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:120]}'), no_update, no_update

    # ── Step 2b: ยกเลิก ──────────────────────────────────────
    @app.callback(
        Output('adm-confirm-section','style', allow_duplicate=True),
        Output('adm-upload-result','children', allow_duplicate=True),
        Output('adm-pending-data','data', allow_duplicate=True),
        Input('adm-btn-cancel','n_clicks'),
        prevent_initial_call=True)
    def cancel_import(n):
        return {'display':'none'},                warn_box('ຍົກເລີກການນຳເຂົ້າແລ້ວ — ບໍ່ມີຂໍ້ມູນຖືກບັນທຶກ'),                None

    # ── Add Student ──────────────────────────────────────────
    @app.callback(Output('adm-add-result','children'), Input('adm-btn-add','n_clicks'),
                  State('adm-code','value'), State('adm-gender','value'), prevent_initial_call=True)
    def add_stu(n, code, gender):
        if not code or not gender: return warn_box('ກະລຸນາປ້ອນຂໍ້ມູນໃຫ້ຄົບ')
        code = code.strip()
        try:
            with db.engine.connect() as conn:
                ex = pd.read_sql(f"SELECT student_id FROM student WHERE student_code='{code}'",conn)
                if len(ex)>0: return warn_box(f'ລະຫັດ "{code}" ມີຢູ່ແລ້ວ')
                nid = int(pd.read_sql("SELECT COALESCE(MAX(student_id),0)+1 AS n FROM student",conn)['n'].iloc[0])
                conn.execute(sa.text("INSERT INTO student(student_id,student_code,gender) VALUES(:i,:c,:g)"),
                             {'i':nid,'c':code,'g':gender})
                conn.commit()
            db.reload_data()
            return ok_box(f'ເພີ່ມ "{code}" ສຳເລັດ', f'{"ຊາຍ" if gender=="M" else "ຍິງ"} · ID: {nid}')
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:80]}')

    # ── Delete Student ───────────────────────────────────────
    @app.callback(Output('adm-del-result','children'), Input('adm-btn-del','n_clicks'),
                  State('adm-del-code','value'), prevent_initial_call=True)
    def del_stu(n, code):
        if not code: return warn_box('ກະລຸນາປ້ອນລະຫັດນັກສຶກສາ')
        code = code.strip()
        try:
            with db.engine.connect() as conn:
                ex = pd.read_sql(f"SELECT student_id FROM student WHERE student_code='{code}'",conn)
                if len(ex)==0: return warn_box(f'ບໍ່ພົບລະຫັດ "{code}" ໃນລະບົບ')
                sid = int(ex['student_id'].iloc[0])
                sc_count = int(pd.read_sql(f"SELECT COUNT(*) AS c FROM score WHERE student_id={sid}",conn)['c'].iloc[0])
                conn.execute(sa.text(f"DELETE FROM score WHERE student_id={sid}"))
                conn.execute(sa.text(f"DELETE FROM student WHERE student_id={sid}"))
                conn.commit()
            db.reload_data()
            return ok_box(f'ລຶບ "{code}" ສຳເລັດ', f'ລຶບຄະແນນທັງໝົດ {sc_count} ແຖວ')
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:80]}')

    # ── Add Subject ──────────────────────────────────────────
    @app.callback(Output('adm-sub-result','children'), Input('adm-btn-sub','n_clicks'),
                  State('adm-subcode','value'), State('adm-subname','value'), prevent_initial_call=True)
    def add_sub(n, code, name):
        if not code or not name: return warn_box('ກະລຸນາປ້ອນຂໍ້ມູນໃຫ້ຄົບ')
        code=code.strip(); name=name.strip()
        try:
            with db.engine.connect() as conn:
                ex = pd.read_sql(f"SELECT subject_id FROM subject WHERE subject_code='{code}'",conn)
                if len(ex)>0: return warn_box(f'ວິຊາ "{code}" ມີຢູ່ແລ້ວ')
                nid = int(pd.read_sql("SELECT COALESCE(MAX(subject_id),0)+1 AS n FROM subject",conn)['n'].iloc[0])
                conn.execute(sa.text("INSERT INTO subject(subject_id,subject_code,subject_name) VALUES(:i,:c,:n)"),
                             {'i':nid,'c':code,'n':name})
                conn.commit()
            db.reload_data()
            return ok_box(f'ເພີ່ມວິຊາ "{code}" ສຳເລັດ', name)
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:80]}')

    # ── Delete Subject ───────────────────────────────────────
    @app.callback(Output('adm-del-sub-result','children'), Input('adm-btn-del-sub','n_clicks'),
                  State('adm-del-subcode','value'), prevent_initial_call=True)
    def del_sub(n, code):
        if not code: return warn_box('ກະລຸນາປ້ອນລະຫັດວິຊາ')
        code = code.strip()
        try:
            with db.engine.connect() as conn:
                ex = pd.read_sql(f"SELECT subject_id FROM subject WHERE subject_code='{code}'",conn)
                if len(ex)==0: return warn_box(f'ບໍ່ພົບວິຊາ "{code}" ໃນລະບົບ')
                bid = int(ex['subject_id'].iloc[0])
                sc_count = int(pd.read_sql(f"SELECT COUNT(*) AS c FROM score WHERE subject_id={bid}",conn)['c'].iloc[0])
                conn.execute(sa.text(f"DELETE FROM score WHERE subject_id={bid}"))
                conn.execute(sa.text(f"DELETE FROM subject WHERE subject_id={bid}"))
                conn.commit()
            db.reload_data()
            return ok_box(f'ລຶບວິຊາ "{code}" ສຳເລັດ', f'ລຶບຄະແນນທັງໝົດ {sc_count} ແຖວ')
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:80]}')

    # ── Edit Grade ───────────────────────────────────────────
    @app.callback(Output('adm-edit-result','children'), Input('adm-btn-edit','n_clicks'),
                  State('adm-edit-stu','value'), State('adm-edit-sub','value'),
                  State('adm-edit-sem','value'), State('adm-edit-grade','value'),
                  prevent_initial_call=True)
    def edit_grade(n, stu_code, sub_code, sem, grade):
        if not all([stu_code, sub_code, sem, grade]): return warn_box('ກະລຸນາປ້ອນຂໍ້ມູນໃຫ້ຄົບທຸກຊ່ອງ')
        gm = {'A':4.0,'B+':3.5,'B':3.0,'C+':2.5,'C':2.0,'D+':1.5,'D':1.0,'F':0.0}
        gp = gm[grade]
        try:
            with db.engine.connect() as conn:
                s = pd.read_sql(f"SELECT student_id FROM student WHERE student_code='{stu_code.strip()}'",conn)
                if len(s)==0: return warn_box(f'ບໍ່ພົບນ.ສ "{stu_code}"')
                b = pd.read_sql(f"SELECT subject_id FROM subject WHERE subject_code='{sub_code.strip()}'",conn)
                if len(b)==0: return warn_box(f'ບໍ່ພົບວິຊາ "{sub_code}"')
                sid = int(s['student_id'].iloc[0]); bid = int(b['subject_id'].iloc[0])
                ex = pd.read_sql(f"SELECT score_id,grade FROM score WHERE student_id={sid} AND subject_id={bid} AND semester='{sem}'",conn)
                if len(ex)==0: return warn_box(f'ບໍ່ພົບຄະແນນ {stu_code} / {sub_code} / {sem}')
                old_grade = ex['grade'].iloc[0]
                conn.execute(sa.text(f"UPDATE score SET grade='{grade}', grade_point={gp} WHERE student_id={sid} AND subject_id={bid} AND semester='{sem}'"))
                conn.commit()
            db.reload_data()
            return ok_box(f'ແກ້ໄຂເກຣດ {stu_code} · {sub_code} · {sem} ສຳເລັດ',
                          f'{old_grade} → {grade} ({gp})')
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:120]}')

    # ── Reset ────────────────────────────────────────────────
    @app.callback(Output('adm-reset-result','children'), Input('adm-btn-reset','n_clicks'),
                  State('adm-reset-confirm','value'), prevent_initial_call=True)
    def reset_all(n, confirm):
        if confirm != 'RESET': return warn_box('ກະລຸນາພິມ "RESET" ໃຫ້ຖືກຕ້ອງເພື່ອຢືນຢັນ')
        try:
            with db.engine.connect() as conn:
                conn.execute(sa.text("DELETE FROM score"))
                conn.execute(sa.text("DELETE FROM student"))
                conn.execute(sa.text("DELETE FROM subject"))
                conn.commit()
            db.reload_data()
            return ok_box('ລ້າງຂໍ້ມູນທັງໝົດສຳເລັດ', 'Student, Subject, Score ຖືກລຶບໝົດແລ້ວ')
        except Exception as e:
            return err_box(f'ຂໍ້ຜິດພາດ: {str(e)[:80]}')